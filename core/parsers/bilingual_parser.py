"""
双语文件解析模块
支持单个文件同时包含中英文本的多种格式
自动识别中文和英文部分
"""

import re
from typing import List, Tuple, Optional, Dict
import logging

logger = logging.getLogger(__name__)


def detect_language(text: str) -> str:
    """
    检测文本语言（支持多语言）
    
    Returns:
        'zh': 中文 (简体/繁体)
        'ja': 日语
        'ko': 韩语
        'ar': 阿拉伯语
        'en': 英语
        'fr': 法语
        'de': 德语
        'ru': 俄语 (西里尔文)
        'latin': 其他拉丁字母语言
        'mixed': 混合或无法判断
    """
    if not text.strip():
        return 'mixed'
    
    # Unicode 字符范围统计
    char_counts = {
        'cjk': 0,           # 中日韩统一表意文字（汉字）
        'hiragana': 0,      # 日语平假名
        'katakana': 0,      # 日语片假名
        'hangul': 0,        # 韩文字母
        'arabic': 0,        # 阿拉伯文
        'cyrillic': 0,      # 西里尔字母（俄语等）
        'latin': 0,         # 拉丁字母（英法德等）
        'greek': 0,         # 希腊字母
        'thai': 0,          # 泰文
    }
    
    # 统计各类字符
    for char in text:
        code = ord(char)
        
        # CJK 统一表意文字 (汉字 - 中文/日文/韩文共用)
        if 0x4E00 <= code <= 0x9FFF or 0x3400 <= code <= 0x4DBF:
            char_counts['cjk'] += 1
        # 日语平假名
        elif 0x3040 <= code <= 0x309F:
            char_counts['hiragana'] += 1
        # 日语片假名
        elif 0x30A0 <= code <= 0x30FF:
            char_counts['katakana'] += 1
        # 韩文字母
        elif 0xAC00 <= code <= 0xD7AF or 0x1100 <= code <= 0x11FF:
            char_counts['hangul'] += 1
        # 阿拉伯文
        elif 0x0600 <= code <= 0x06FF or 0x0750 <= code <= 0x077F:
            char_counts['arabic'] += 1
        # 西里尔字母（俄语、乌克兰语等）
        elif 0x0400 <= code <= 0x04FF:
            char_counts['cyrillic'] += 1
        # 希腊字母
        elif 0x0370 <= code <= 0x03FF:
            char_counts['greek'] += 1
        # 泰文
        elif 0x0E00 <= code <= 0x0E7F:
            char_counts['thai'] += 1
        # 拉丁字母（A-Z, a-z, 以及扩展如 é, ö, β 等）
        elif (0x0041 <= code <= 0x005A or 0x0061 <= code <= 0x007A or 
              0x00C0 <= code <= 0x00FF or 0x0100 <= code <= 0x017F or 0x00DF == code):
            char_counts['latin'] += 1
    
    total_chars = sum(char_counts.values())
    
    if total_chars == 0:
        return 'mixed'
    
    # 语言占比
    ratios = {k: v / total_chars for k, v in char_counts.items()}
    
    # 判断规则（按优先级）
    
    # 1. 日语：有平假名或片假名（即使混有汉字）
    if char_counts['hiragana'] + char_counts['katakana'] >= 1:
        return 'ja'
    
    # 2. 韩语：有韩文字母
    if char_counts['hangul'] >= 1:
        return 'ko'
    
    # 3. 阿拉伯语：有阿拉伯文字符
    if char_counts['arabic'] >= 1:
        return 'ar'
    
    # 4. 西里尔字母（俄语等）
    if char_counts['cyrillic'] >= 2:
        return 'ru'
    
    # 5. 中文 vs 日语 (极致对比逻辑)
    if char_counts['cjk'] >= 1:
        # 如果有假名，坚决认定为日语
        if char_counts['hiragana'] + char_counts['katakana'] >= 1:
            return 'ja'
            
        # 汉字对比：寻找日语独有汉字 (Kokuji) vs 简体中文特征字
        ja_kanji_only = {
            '込', '枠', '峠', '畑', '駅', '込', '雫', '喰', '峠', '腺', '栃', '匂', '塀', '搾', '吊', '噛', '揃', '蹴'
        }
        # 简体中文高度特征字 (日语中不使用的字形)
        sc_chars_only = {
            '们', '这', '说', '书', '会', '为', '这', '国', '学', '体', '时', '应', '对', '说', '发', '现', '经', '济'
        }
        
        text_chars = set(text)
        
        # 检查是否包含日语专用汉字
        if any(c in text_chars for c in ja_kanji_only):
            return 'ja'
            
        # 检查是否包含简体中文独有特征字
        if any(c in text_chars for c in sc_chars_only):
            return 'zh'
            
        # 如果是繁体字且没有日语特征，暂时回退到 zh（涵盖繁体中文）
        return 'zh'
    
    # 6. 拉丁字母语言（英法德等）
    if char_counts['latin'] >= 1:
        # 预判断：如果有特定字符
        text_lower = text.lower()
        
        # 德语特有字符 (ß, ä, ö, ü)
        if any(c in text_lower for c in 'ßäöü'):
            return 'de'
        # 法语特有字符 (é, à, è, ù, â, ê, î, ô, û, ë, ï, ç)
        if any(c in text_lower for c in 'éàèùâêîôûëïç'):
            return 'fr'
        
        # 简单启发式：检测常见限定词和连词 (需完整匹配单词)
        words = set(re.findall(r'\b\w+\b', text_lower))
        
        # 法语特征词
        french_markers = {
            'le', 'la', 'les', 'un', 'une', 'des', 'du', 'de', 'et', 'est', 'être', 
            'je', 'vous', 'pour', 'dans', 'qui', 'que', 'en', 'par', 'ce', 'pas', 'sur', 'bonjour'
        }
        # 德语特征词  
        german_markers = {
            'der', 'die', 'das', 'den', 'dem', 'des', 'und', 'ist', 'sind', 'ich', 'sie',
            'nicht', 'mit', 'zu', 'für', 'auf', 'ein', 'eine', 'guten', 'tag', 'danke', 'ja', 'nein'
        }
        # 英语特征词
        english_markers = {
            'the', 'a', 'an', 'and', 'is', 'are', 'in', 'of', 'to', 'it', 'you', 'that', 
            'for', 'with', 'on', 'at', 'by', 'this', 'have', 'from', 'hello', 'thank'
        }
        
        fr_score = sum(1 for w in words if w in french_markers)
        de_score = sum(1 for w in words if w in german_markers)
        en_score = sum(1 for w in words if w in english_markers)
        
        # 记录特定语言分数
        if fr_score > de_score and fr_score > en_score:
            return 'fr'
        elif de_score > fr_score and de_score > en_score:
            return 'de'
        elif en_score > fr_score and en_score > de_score:
            return 'en'
            
        # 如果得分相同或全为0，尝试通过字母频率或特殊重音
        if any(c in text_lower for c in 'éàèù'): return 'fr'
        if any(c in text_lower for c in 'äöüß'): return 'de'
        
        # 默认英语 (对于大多数拉丁字母文本)
        return 'en'
    
    # 7. 希腊语
    if char_counts['greek'] >= 1:
        return 'greek'
    
    # 8. 泰语
    if char_counts['thai'] >= 1:
        return 'thai'
    
    return 'mixed'


def smart_language_pair(text1: str, text2: str) -> Tuple[str, str]:
    """
    智能识别两段文本的语言，并返回 (source, target)
    支持多语言组合
    
    Returns:
        (source_text, target_text) - 根据语言自动分配
    """
    lang1 = detect_language(text1)
    lang2 = detect_language(text2)
    
    # 语言分组（常见的原文语言 vs 目标语言）
    # 通常：英语/日语 作为原文，中文/其他语言作为译文
    source_langs = {'en', 'ja', 'fr', 'de', 'ru', 'ko'}  # 常见原文语言
    target_langs = {'zh'}  # 常见目标语言（中文）
    
    # 规则1: 如果一个是常见原文语言，一个是中文，按原文→中文排列
    if lang1 in source_langs and lang2 == 'zh':
        return (text1, text2)
    elif lang2 in source_langs and lang1 == 'zh':
        return (text2, text1)
    
    # 规则2: 如果都是非中文，保持原顺序
    # 或者都是中文，保持原顺序
    # 或者一个是mixed，继承顺序
    if lang1 == lang2 or lang1 == 'mixed' or lang2 == 'mixed':
        logger.debug(f"语言相同或混合: {lang1} vs {lang2}，保持原顺序")
        return (text1, text2)
    
    # 规则3: 其他语言组合，按字典序排列（保证一致性）
    if lang1 < lang2:
        return (text1, text2)
    else:
        return (text2, text1)



def detect_bilingual_format(file_path: str) -> str:
    """
    自动检测双语文件格式
    
    Returns:
        'alternating': 行交替格式 (偶数行英文，奇数行中文)
        'delimiter': 分隔符格式 (|, \t 等)
        'tsv': TSV/CSV 格式
        'block': 块状格式 (前半部分英文，后半部分中文)
        'unknown': 无法识别
    """
    try:
        # 首先检查文件扩展名 - 如果是标准字幕格式，直接返回unknown
        # 标准字幕文件不应该被当作双语文件
        ext = file_path.lower().split('.')[-1]
        if ext in ['srt', 'ass', 'ssa', 'vtt']:
            logger.debug(f"跳过标准字幕文件的双语检测: {ext}")
            return 'unknown'
        
        # Excel 文件检测
        if ext in ['xlsx', 'xls']:
            return 'xlsx'
        
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = [line.strip() for line in f.readlines() if line.strip()]
            
        if not lines or len(lines) < 4:
            return 'unknown'
        
        # 检测 TSV/CSV (第一行包含表头)
        if '\t' in lines[0] and any(keyword in lines[0].lower() for keyword in ['en', 'cn', 'source', 'target']):
            return 'tsv'
            
        # 检测分隔符格式
        delimiters = ['|', '\t', '|||']
        for delimiter in delimiters:
            if delimiter in lines[0]:
                # 验证多行都有分隔符
                count = sum(1 for line in lines if delimiter in line)
                if count >= len(lines) * 0.8:  # 80%的行都有分隔符
                    return 'delimiter'
        
        # 检测行交替格式 - 需要更严格的条件
        # 要求：偶数行必须主要是拉丁字符，奇数行必须主要是CJK字符
        even_lines_latin = 0
        odd_lines_cjk = 0
        total_checkable = min(20, len(lines) // 2 * 2)  # 只检查偶数数量的行
        
        for i in range(total_checkable):
            line = lines[i]
            # 统计各类字符
            latin_chars = len(re.findall(r'[a-zA-Z]', line))
            cjk_chars = len(re.findall(r'[\u4e00-\u9fff]', line))
            total_chars = len(re.findall(r'[\w\u4e00-\u9fff]', line))
            
            if total_chars < 5:  # 太短的行跳过
                continue
            
            if i % 2 == 0:  # 偶数行应该是英文
                if latin_chars > cjk_chars and latin_chars / total_chars > 0.6:
                    even_lines_latin += 1
            else:  # 奇数行应该是中文
                if cjk_chars > latin_chars and cjk_chars / total_chars > 0.6:
                    odd_lines_cjk += 1
        
        # 要求至少60%的偶数行是英文，60%的奇数行是中文
        half_checkable = total_checkable // 2
        if (even_lines_latin >= half_checkable * 0.6 and 
            odd_lines_cjk >= half_checkable * 0.6):
            logger.info(f"检测到交替格式: {even_lines_latin}/{half_checkable} 英文行, {odd_lines_cjk}/{half_checkable} 中文行")
            return 'alternating'
        
        # 检测块状格式 (前半部分一种语言，后半部分另一种语言)
        # 策略：检查前40%和后40%的语言分布
        total_lines = len(lines)
        if total_lines < 20:  # 太少的行不检测块状格式
            return 'unknown'
        
        first_part_end = int(total_lines * 0.4)
        second_part_start = int(total_lines * 0.6)
        
        first_part_lines = lines[:first_part_end]
        second_part_lines = lines[second_part_start:]
        
        # 统计前40%的语言
        first_part_lang = _detect_text_block_language(first_part_lines)
        # 统计后40%的语言
        second_part_lang = _detect_text_block_language(second_part_lines)
        
        # 如果前后两部分语言不同，判定为块状格式
        if first_part_lang != 'mixed' and second_part_lang != 'mixed' and first_part_lang != second_part_lang:
            logger.info(f"检测到块状格式: 前部分={first_part_lang}, 后部分={second_part_lang}")
            return 'block'
            
        return 'unknown'
        
    except Exception as e:
        logger.error(f"检测双语格式失败: {e}")
        return 'unknown'


def _detect_text_block_language(lines: List[str]) -> str:
    """
    检测一组文本行的主要语言
    
    Returns:
        'en', 'zh', or 'mixed'
    """
    total_cjk = 0
    total_latin = 0
    
    for line in lines:
        total_cjk += len(re.findall(r'[\u4e00-\u9fff]', line))
        total_latin += len(re.findall(r'[a-zA-Z]', line))
    
    total = total_cjk + total_latin
    
    if total == 0:
        return 'mixed'
    
    if total_cjk / total > 0.6:
        return 'zh'
    elif total_latin / total > 0.6:
        return 'en'
    else:
        return 'mixed'



def parse_bilingual_file(file_path: str, format_hint: Optional[str] = None, 
                         smart_detect: bool = True,
                         use_llm_alignment: bool = False,
                         api_client=None) -> List[Tuple[str, str]]:
    """
    解析双语文件，返回 (原文, 译文) 对列表
    
    Args:
        file_path: 文件路径
        format_hint: 格式提示 (可选)，如果不提供则自动检测
        smart_detect: 是否使用智能语言检测自动识别中英文 (默认 True)
        use_llm_alignment: 当块状格式行数不等时，是否使用 LLM 智能对齐
        api_client: API 客户端（LLM 对齐时需要）
        
    Returns:
        [(source, target), ...] 列表，自动识别并排列为 英文→中文
    """
    if format_hint is None:
        format_hint = detect_bilingual_format(file_path)
        logger.info(f"检测到双语格式: {format_hint}")
    
    try:
        if format_hint == 'alternating':
            raw_pairs = _parse_alternating(file_path)
        elif format_hint == 'block':
            raw_pairs = _parse_block(file_path, use_llm_alignment, api_client)
        elif format_hint in ['delimiter', 'tsv']:
            raw_pairs = _parse_delimiter(file_path)
        elif format_hint == 'xlsx':
            raw_pairs = parse_xlsx_bilingual(file_path)
        else:
            logger.warning(f"未知的双语格式，尝试自动解析")
            # 尝试分隔符格式
            result = _parse_delimiter(file_path)
            if result:
                raw_pairs = result
            else:
                # 尝试行交替格式
                raw_pairs = _parse_alternating(file_path)
        
        # 如果启用智能检测，自动识别并排列为 source → target
        if smart_detect and raw_pairs:
            smart_pairs = []
            for text1, text2 in raw_pairs:
                source, target = smart_language_pair(text1, text2)
                smart_pairs.append((source, target))
            
            logger.info(f"智能语言检测完成: {len(smart_pairs)} 对")
            return smart_pairs
        else:
            return raw_pairs
            
    except Exception as e:
        logger.error(f"解析双语文件失败: {e}")
        return []


def _parse_alternating(file_path: str) -> List[Tuple[str, str]]:
    """解析行交替格式"""
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]
    
    result = []
    for i in range(0, len(lines) - 1, 2):
        source = lines[i]
        target = lines[i + 1] if i + 1 < len(lines) else ""
        result.append((source, target))
    
    logger.info(f"行交替格式解析完成: {len(result)} 对")
    return result


def _parse_block(file_path: str, use_llm_alignment: bool = False, 
                 api_client=None) -> List[Tuple[str, str]]:
    """
    解析块状格式 (前半部分英文，后半部分中文)
    
    策略：
    1. 逐行识别语言，分组成语言块
    2. 如果英文和中文行数相等 → 直接1:1配对
    3. 如果行数不等 → 使用 LLM 智能对齐（需要提供 api_client）
    
    Args:
        file_path: 文件路径
        use_llm_alignment: 当行数不等时是否使用 LLM 对齐
        api_client: API 客户端（LLM 对齐时需要）
    """
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = [line.strip() for line in f.readlines() if line.strip()]
    
    total_lines = len(lines)
    
    if total_lines < 4:
        logger.warning("文件行数太少，无法解析块状格式")
        return []
    
    # 步骤1: 逐行检测语言并分组
    line_languages = [detect_language(line) for line in lines]
    
    # 识别语言块
    blocks = []
    current_block_lang = None
    current_block_start = 0
    
    for i, lang in enumerate(line_languages):
        if lang == 'mixed':
            # 混合语言，继承前一行的语言
            if i > 0:
                lang = line_languages[i-1]
            elif i < len(line_languages) - 1:
                lang = line_languages[i+1]
        
        if current_block_lang is None:
            current_block_lang = lang
            current_block_start = i
        elif lang != current_block_lang:
            blocks.append({
                'lang': current_block_lang,
                'start': current_block_start,
                'end': i,
                'lines': lines[current_block_start:i]
            })
            current_block_lang = lang
            current_block_start = i
    
    # 保存最后一个块
    if current_block_lang is not None:
        blocks.append({
            'lang': current_block_lang,
            'start': current_block_start,
            'end': len(lines),
            'lines': lines[current_block_start:]
        })
    
    logger.info(f"识别到 {len(blocks)} 个语言块")
    for i, block in enumerate(blocks):
        logger.info(f"  块 {i+1}: {block['lang']} ({len(block['lines'])} 行)")
    
    # 步骤2: 提取并合并所有英文块和中文块
    en_blocks = [b for b in blocks if b['lang'] == 'en']
    zh_blocks = [b for b in blocks if b['lang'] == 'zh']
    
    if not en_blocks or not zh_blocks:
        logger.warning(f"未找到有效的英中文块: 英文块={len(en_blocks)}, 中文块={len(zh_blocks)}")
        return []
    
    # 合并所有英文行（保持顺序）
    en_lines = []
    for block in en_blocks:
        en_lines.extend(block['lines'])
    
    # 合并所有中文行（保持顺序）
    zh_lines = []
    for block in zh_blocks:
        zh_lines.extend(block['lines'])
    
    logger.info(f"合并后: 英文 {len(en_lines)} 行 (来自 {len(en_blocks)} 个块), 中文 {len(zh_lines)} 行 (来自 {len(zh_blocks)} 个块)")
    
    # 步骤3: 根据行数决定配对策略
    if len(en_lines) == len(zh_lines):
        # 行数相等，直接1:1配对
        logger.info("✅ 行数相等，执行直接配对")
        result = [(en_lines[i], zh_lines[i]) for i in range(len(en_lines))]
        logger.info(f"块状格式解析完成: {len(result)} 对")
        return result
    
    else:
        # 行数不等
        logger.warning(f"⚠️ 行数不等: 英文 {len(en_lines)} 行 vs 中文 {len(zh_lines)} 行")
        
        if use_llm_alignment and api_client:
            # 使用 LLM 智能对齐
            logger.info("使用 LLM 智能对齐...")
            return _align_with_llm(en_lines, zh_lines, api_client)
        else:
            # 简单截断配对，并警告
            result = []
            for i in range(min(len(en_lines), len(zh_lines))):
                result.append((en_lines[i], zh_lines[i]))
            
            logger.warning(f"⚠️ 未使用 LLM 对齐，采用简单配对: {len(result)} 对")
            logger.warning(f"提示: 使用 parse_bilingual_file(..., use_llm_alignment=True, api_client=client) 进行智能对齐")
            return result


def _align_with_llm(en_lines: List[str], zh_lines: List[str], api_client) -> List[Tuple[str, str]]:
    """
    使用 LLM 智能对齐不等长的英中文本
    
    Args:
        en_lines: 英文行列表
        zh_lines: 中文行列表
        api_client: API 客户端
        
    Returns:
        对齐后的 (英文, 中文) 对列表
    """
    logger.info("开始 LLM 智能对齐...")
    
    # 构建 Prompt
    system_prompt = """你是一个专业的双语文本对齐专家。

任务：给定一组英文句子和一组中文翻译句子，需要将它们正确配对。

规则：
1. 保持句子的原始顺序，不能打乱
2. 可以合并多个句子对应一个翻译（例如：英文3句 → 中文1句）
3. 如果某个英文句子没有对应的中文翻译，中文部分填空字符串 ""
4. 如果某个中文句子没有对应的英文原文，英文部分填空字符串 ""

输出格式：JSON数组，每个元素包含：
{
  "en": "英文句子（可以是多句合并）",
  "zh": "对应的中文翻译"
}

确保输出是有效的 JSON 格式，不要包含 markdown 代码块标记。"""
    
    user_prompt = f"""英文句子 ({len(en_lines)} 行):
{chr(10).join([f"{i+1}. {line}" for i, line in enumerate(en_lines)])}

中文句子 ({len(zh_lines)} 行):
{chr(10).join([f"{i+1}. {line}" for i, line in enumerate(zh_lines)])}

请将它们正确配对，输出 JSON 数组。"""
    
    try:
        response = api_client.generate_content(
            system_prompt=system_prompt,
            user_prompt=user_prompt,
            json_mode=True,
            temperature=0.3  # 低温度，确保一致性
        )
        
        # 解析响应
        result_text = response['text']
        
        # 处理可能的 Markdown 代码块包装
        if result_text.strip().startswith('```'):
            lines = result_text.strip().split('\n')
            result_text = '\n'.join(lines[1:-1])
        
        import json
        pairs_json = json.loads(result_text)
        
        # 转换为元组列表
        result = []
        for item in pairs_json:
            en = item.get('en', '')
            zh = item.get('zh', '')
            result.append((en, zh))
        
        logger.info(f"✅ LLM 对齐完成: {len(result)} 对")
        return result
        
    except Exception as e:
        logger.error(f"LLM 对齐失败: {e}")
        logger.warning("回退到简单配对")
        
        # 回退：简单配对
        result = []
        for i in range(min(len(en_lines), len(zh_lines))):
            result.append((en_lines[i], zh_lines[i]))
        return result


def _parse_delimiter(file_path: str) -> List[Tuple[str, str]]:
    """解析分隔符格式 (|, \t 等)"""
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    # 检测分隔符
    delimiter = '\t'
    if '|' in lines[0]:
        delimiter = '|'
    elif '|||' in lines[0]:
        delimiter = '|||'
    
    result = []
    skip_header = False
    
    # 检查是否有表头
    first_line_parts = lines[0].strip().split(delimiter)
    if any(keyword in first_line_parts[0].lower() for keyword in ['en', 'source', 'english']):
        skip_header = True
    
    start_idx = 1 if skip_header else 0
    
    for line in lines[start_idx:]:
        parts = line.strip().split(delimiter)
        if len(parts) >= 2:
            source = parts[0].strip()
            target = parts[1].strip()
            if source or target:  # 至少有一个不为空
                result.append((source, target))
    
    logger.info(f"分隔符格式解析完成: {len(result)} 对 (分隔符: '{delimiter}')")
    return result


def is_bilingual_file(file_path: str) -> bool:
    """
    判断文件是否为双语文件
    
    Returns:
        True if file contains bilingual content
    """
    format_type = detect_bilingual_format(file_path)
    return format_type != 'unknown'


def get_language_stats(pairs: List[Tuple[str, str]]) -> Dict[str, int]:
    """
    获取双语对的语言统计信息
    
    Args:
        pairs: [(text1, text2), ...] 列表
        
    Returns:
        统计信息字典
    """
    stats = {
        'total': len(pairs),
        'source_en': 0,
        'source_zh': 0,
        'source_mixed': 0,
        'target_en': 0,
        'target_zh': 0,
        'target_mixed': 0
    }
    
    for source, target in pairs:
        source_lang = detect_language(source)
        target_lang = detect_language(target)
        
        if source_lang == 'en':
            stats['source_en'] += 1
        elif source_lang == 'zh':
            stats['source_zh'] += 1
        else:
            stats['source_mixed'] += 1
        
        if target_lang == 'en':
            stats['target_en'] += 1
        elif target_lang == 'zh':
            stats['target_zh'] += 1
        else:
            stats['target_mixed'] += 1
    
    return stats


def parse_xlsx_bilingual(file_path: str) -> List[Tuple[str, str]]:
    """
    解析 Excel (.xlsx) 双语文件
    
    支持的格式：
    1. 两列格式：A列原文，B列译文（可选表头）
    2. 多列格式：包含 source/target, en/zh, 原文/译文 等标识列
    
    Args:
        file_path: Excel 文件路径
        
    Returns:
        [(source, target), ...] 双语对列表
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("未安装 openpyxl 库，无法解析 Excel 文件")
        logger.error("请运行: pip install openpyxl")
        return []
    
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            logger.warning("Excel 文件为空")
            return []
        
        # 检查第一行是否为表头
        first_row = rows[0]
        has_header = False
        source_col_idx = 0
        target_col_idx = 1
        
        # 表头识别关键词
        if first_row and len(first_row) >= 2:
            col1_text = str(first_row[0] or "").lower()
            col2_text = str(first_row[1] or "").lower()
            
            # 检测表头
            header_keywords = ['source', 'target', 'en', 'zh', 'english', 'chinese', '原文', '译文', '英文', '中文']
            if any(kw in col1_text or kw in col2_text for kw in header_keywords):
                has_header = True
                logger.info(f"检测到表头: {first_row[0]} | {first_row[1]}")
                
                # 智能识别列顺序
                source_keywords = ['source', 'en', 'english', '原文', '英文']
                target_keywords = ['target', 'zh', 'chinese', '译文', '中文']
                
                if any(kw in col1_text for kw in target_keywords) and any(kw in col2_text for kw in source_keywords):
                    # 列顺序颠倒
                    source_col_idx = 1
                    target_col_idx = 0
                    logger.info("检测到列顺序：B列为原文，A列为译文")
        
        # 解析数据行
        start_row = 1 if has_header else 0
        result = []
        
        for i, row in enumerate(rows[start_row:], start=start_row):
            if not row or len(row) < 2:
                continue
            
            source_text = str(row[source_col_idx] or "").strip()
            target_text = str(row[target_col_idx] or "").strip()
            
            # 跳过空行
            if not source_text and not target_text:
                continue
            
            # 智能语言判断
            source_final, target_final = smart_language_pair(source_text, target_text)
            result.append((source_final, target_final))
        
        logger.info(f"Excel 解析完成: {len(result)} 对")
        wb.close()
        return result
        
    except Exception as e:
        logger.error(f"解析 Excel 文件失败: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return []


