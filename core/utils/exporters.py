
"""
数据导出模块
负责将字幕和LQA数据导出为不同格式
"""
import os
import json
import csv
import logging
try:
    import pysubs2
except ImportError:
    pysubs2 = None

logger = logging.getLogger(__name__)

class DataExporter:
    """数据导出工具类"""
    
    @staticmethod
    def export_json_report(subtitle_data, file_path):
        """导出JSON格式LQA报告"""
        try:
            report_data = []
            for i, item in enumerate(subtitle_data):
                src = item.get('source', {})
                tgt = item.get('target', {})
                lqa_result = item.get('lqa_result')
                
                src_text = src.get('text', '') if isinstance(src, dict) else str(src)
                tgt_text = tgt.get('text', '') if isinstance(tgt, dict) else str(tgt)
                
                report_data.append({
                    'id': i + 1,
                    'start': src.get('start', 0) if isinstance(src, dict) else 0,
                    'end': src.get('end', 0) if isinstance(src, dict) else 0,
                    'source': src_text,
                    'target': tgt_text,
                    'score': lqa_result.get('score', 0) if lqa_result else 0,
                    'issues': lqa_result.get('issues', []) if lqa_result else [],
                    'suggestions': lqa_result.get('suggestions', '') if lqa_result else ''
                })
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(report_data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"已导出JSON报告: {file_path}")
            return True, f"已导出JSON报告: {os.path.basename(file_path)}"
            
        except Exception as e:
            logger.error(f"导出JSON失败: {e}")
            return False, str(e)

    @staticmethod
    def export_suggestions(subtitle_data, file_path, time_base='source'):
        """导出LQA建议译文"""
        try:
            suggestions = []
            for i, item in enumerate(subtitle_data):
                src = item.get('source', {})
                tgt = item.get('target', {})
                lqa_result = item.get('lqa_result')
                
                # 优先使用LQA建议，否则使用原译文
                if lqa_result and lqa_result.get('suggestions'):
                    text = lqa_result.get('suggestions', '')
                else:
                    text = tgt.get('text', '') if isinstance(tgt, dict) else str(tgt)
                
                # 获取时间戳
                time_obj = src if time_base == 'source' else tgt
                start = time_obj.get('start', 0) if isinstance(time_obj, dict) else 0
                end = time_obj.get('end', 0) if isinstance(time_obj, dict) else 0
                
                suggestions.append({
                    'id': i + 1,
                    'start': start,
                    'end': end,
                    'text': text
                })
            
            if file_path.endswith('.srt'):
                DataExporter._write_srt(file_path, suggestions)
            elif file_path.endswith('.json'):
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(suggestions, f, indent=2, ensure_ascii=False)
            elif file_path.lower().endswith(('.ass', '.ssa', '.vtt')):
                DataExporter._save_via_pysubs2(suggestions, file_path)
            else:  # .txt
                with open(file_path, 'w', encoding='utf-8') as f:
                    for item in suggestions:
                        f.write(f"{item['text']}\n")
                        
            logger.info(f"已导出建议译文: {file_path}")
            return True, f"已导出 {len(suggestions)} 条建议译文"
            
        except Exception as e:
            logger.error(f"导出建议译文失败: {e}")
            return False, str(e)

    @staticmethod
    def _write_srt(file_path, subtitles):
        """写入SRT格式"""
        def format_timestamp(ms):
            if ms == 0: return "00:00:00,000"
            hours = ms // 3600000
            ms %= 3600000
            minutes = ms // 60000
            ms %= 60000
            seconds = ms // 1000
            milliseconds = ms % 1000
            return f"{hours:02d}:{minutes:02d}:{seconds:02d},{milliseconds:03d}"

        with open(file_path, 'w', encoding='utf-8') as f:
            for item in subtitles:
                f.write(f"{item['id']}\n")
                start_ms = int(item['start'] * 1000)
                end_ms = int(item['end'] * 1000)
                f.write(f"{format_timestamp(start_ms)} --> {format_timestamp(end_ms)}\n")
                f.write(f"{item['text']}\n\n")

    @staticmethod
    def export_content(subtitle_data, file_path, side='target', time_base='source'):
        """导出纯文本或字幕（原文或译文）"""
        try:
            items = []
            for i, item in enumerate(subtitle_data):
                source = item.get('source', {})
                target = item.get('target', {})
                
                # 选择侧
                data_obj = source if side == 'source' else target
                # 获取文本，确保 safely get text
                text = data_obj.get('text', '') if isinstance(data_obj, dict) else str(data_obj)
                
                # 获取时间戳（根据 time_base）
                time_obj = source if time_base == 'source' else target
                start = time_obj.get('start', 0) if isinstance(time_obj, dict) else 0
                end = time_obj.get('end', 0) if isinstance(time_obj, dict) else 0
                
                items.append({
                    'id': i + 1,
                    'start': start,
                    'end': end,
                    'text': text
                })
            
            if file_path.lower().endswith('.srt'):
                DataExporter._write_srt(file_path, items)
            elif file_path.lower().endswith('.json'):
                with open(file_path, 'w', encoding='utf-8') as f:
                    json.dump(items, f, indent=2, ensure_ascii=False)
            elif file_path.lower().endswith(('.ass', '.ssa', '.vtt')):
                DataExporter._save_via_pysubs2(items, file_path)
            else:  # .txt 等
                with open(file_path, 'w', encoding='utf-8') as f:
                    for item in items:
                        f.write(f"{item['text']}\n")
                        
            logger.info(f"已导出 {side}: {file_path}")
            return True, f"已导出 {len(items)} 条{side}内容"
            
        except Exception as e:
            logger.error(f"导出 {side} 失败: {e}")
            return False, str(e)

    @staticmethod
    def export_csv(subtitle_data, file_path):
        """导出CSV格式"""
        try:
            # Helper for timestamp formatting
            def format_timestamp(ms):
                if ms == 0: return "00:00:00.000"
                hours = ms // 3600000
                ms %= 3600000
                minutes = ms // 60000
                ms %= 60000
                seconds = ms // 1000
                milliseconds = ms % 1000
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}.{milliseconds:03d}"

            with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['ID', '开始', '结束', '原文', '译文', '评分', '问题', '建议'])
                
                for i, item in enumerate(subtitle_data):
                    src = item.get('source', {})
                    tgt = item.get('target', {})
                    lqa_result = item.get('lqa_result')
                    
                    src_text = src.get('text', '') if isinstance(src, dict) else str(src)
                    tgt_text = tgt.get('text', '') if isinstance(tgt, dict) else str(tgt)
                    
                    start_ms = int(src.get('start', 0) * 1000) if isinstance(src, dict) else 0
                    end_ms = int(src.get('end', 0) * 1000) if isinstance(src, dict) else 0
                    
                    start_str = format_timestamp(start_ms) if start_ms else ''
                    end_str = format_timestamp(end_ms) if end_ms else ''
                    
                    score = lqa_result.get('score', '') if lqa_result else ''
                    issues = ', '.join(lqa_result.get('issues', [])) if lqa_result else ''
                    suggestions = lqa_result.get('suggestions', '') if lqa_result else ''
                    
                    writer.writerow([i+1, start_str, end_str, src_text, tgt_text, score, issues, suggestions])
            
            logger.info(f"已导出CSV: {file_path}")
            return True, f"已导出CSV: {os.path.basename(file_path)}"
            
        except Exception as e:
            logger.error(f"导出CSV失败: {e}")
            return False, str(e)

    @staticmethod
    def _save_via_pysubs2(items, file_path):
        """使用pysubs2保存高级字幕格式 (ASS, VTT)"""
        if not pysubs2:
            raise ImportError("未安装 pysubs2 库，无法导出 ASS/VTT 格式")
            
        subs = pysubs2.SSAFile()
        for item in items:
            start_ms = int(item['start'] * 1000)
            end_ms = int(item['end'] * 1000)
            text = item['text']
            event = pysubs2.SSAEvent(start=start_ms, end=end_ms, text=text)
            subs.events.append(event)
            
        subs.save(file_path, encoding='utf-8')

    @staticmethod
    def export_xlsx_report(subtitle_data, file_path):
        """导出 Excel 格式 LQA 报告"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            logger.error("未安装 openpyxl 库，无法导出 Excel")
            return False, "需要安装 openpyxl: pip install openpyxl"
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "LQA报告"
            
            # 表头
            headers = ['序号', '开始时间', '结束时间', '原文', '译文', '评分', '问题', '建议']
            ws.append(headers)
            
            # 设置表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 数据行
            for i, item in enumerate(subtitle_data, start=1):
                src = item.get('source', {})
                tgt = item.get('target', {})
                lqa_result = item.get('lqa_result')
                
                src_text = src.get('text', '') if isinstance(src, dict) else str(src)
                tgt_text = tgt.get('text', '') if isinstance(tgt, dict) else str(tgt)
                
                start = src.get('start', 0) if isinstance(src, dict) else 0
                end = src.get('end', 0) if isinstance(src, dict) else 0
                
                # 格式化时间
                start_str = DataExporter._format_time(start) if start else ''
                end_str = DataExporter._format_time(end) if end else ''
                
                score = lqa_result.get('score', '') if lqa_result else ''
                issues = ', '.join(lqa_result.get('issues', [])) if lqa_result else ''
                suggestions = lqa_result.get('suggestions', '') if lqa_result else ''
                
                row = [i, start_str, end_str, src_text, tgt_text, score, issues, suggestions]
                ws.append(row)
            
            # 调整列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 40
            ws.column_dimensions['F'].width = 8
            ws.column_dimensions['G'].width = 30
            ws.column_dimensions['H'].width = 40
            
            wb.save(file_path)
            logger.info(f"已导出 Excel 报告: {file_path}")
            return True, f"已导出 Excel 报告: {os.path.basename(file_path)}"
            
        except Exception as e:
            logger.error(f"导出 Excel 失败: {e}")
            return False, str(e)
    
    @staticmethod
    def _format_time(seconds):
        """将秒数格式化为 HH:MM:SS.mmm"""
        if seconds == 0:
            return "00:00:00.000"
        hours = int(seconds // 3600)
        seconds %= 3600
        minutes = int(seconds // 60)
        seconds %= 60
        return f"{hours:02d}:{minutes:02d}:{seconds:06.3f}"

